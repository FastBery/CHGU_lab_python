import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('./15_lab/18.xlsx')

sheet = wb.active
try:
    wb['2']
except:
    wb.create_sheet('2')
    wb.save('./15_lab/18.xlsx')

sheet2 = wb["2"]


columns = sheet.max_row

n = 0

s = Side(border_style="double")

for i in range(1, columns + 1):
    mul = 1
    sum = 0
    array = []
    # print(sheet[i])
    for j in sheet[i][:3]:
        if j.value != None:
            array.append(j.value)
            for k in str(j.value):
                mul *= int(k)
            sum += j.value
    sheet[f"D{i}"] = str(sum)
    sheet[f"E{i}"] = str(mul)
    if sum < mul:
        n += 1
        sheet2.append(array)
        for j in sheet[i]:
            j.fill = PatternFill("solid", fgColor="DDDDDD")
            j.border = Border(s,s,s,s)

# print(sheet[f"A{i}:C{i}"])
        



# print(sheet.rows[1])
# for i in sheet.rows:
#     print(i)
#     mul = 1
#     for j in i:
#         for k in str(j.value()):
#             mul *= int(k)
    
    # print(mul)


# for i in sheet.iter_rows(values_only=True):
#     mul = 1
#     for k in i:
#         for j in str(k):
#             mul *= int(j)
    
#     if sum(i) < mul:
#         n += 1
#         sheet2.append(i)

wb.save('./15_lab/18.xlsx')
print(n)