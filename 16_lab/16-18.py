import sqlite3 as sq
import openpyxl

db = sq.connect('./16_lab/data_base/18.db')
cur = db.cursor()

cur.execute("PRAGMA foreign_keys=ON;") #making foreign_keys=on

cur.execute("""CREATE TABLE IF NOT EXISTS Services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            price INTEGER NOT NULL
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            district TEXT NOT NULL,
            address TEXT NOT NULL,
            name TEXT NOT NULL
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Rendered (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            client_id INTEGER NOT NULL,
            service_id INTEGER NOT NULL,
            FOREIGN KEY (client_id) REFERENCES Clients (id) ON DELETE CASCADE,
            FOREIGN KEY (service_id) REFERENCES Services (id) ON DELETE CASCADE
            )""")

wb = openpyxl.load_workbook('./16_lab/18_original.xlsx')

clients_sheet = wb["Клиенты"]
services_sheet = wb["Услуги"]
rendered_sheet = wb["Оказанные услуги"]

#data for clients
clients_data = []
for i in clients_sheet.iter_rows(values_only=True):
    clients_data.append(i)

clients_data.pop(0)

try:
    cur.executemany("""INSERT INTO Clients (id, district, address, name) VALUES (?, ?, ?, ?)""", clients_data)
except:
    pass


#data for services
services_data = []
for i in services_sheet.iter_rows(values_only=True):
    services_data.append(i)

services_data.pop(0)

try:
    cur.executemany("""INSERT INTO Services (id, name, price) VALUES (?, ?, ?)""", services_data)
except:
    pass


#data for renders
rendered_data = []
for i in rendered_sheet.iter_rows(values_only=True):
    rendered_data.append(i)

rendered_data.pop(0)

try:
    cur.executemany("""INSERT INTO Rendered (id, date, client_id, service_id) VALUES (?, ?, ?, ?)""", rendered_data)
except:
    pass

# cur.execute("""SELECT * FROM Rendered""")
# for i in cur.fetchall():
#     print(i)


def date_iter(d1, d2, month, year):
    if month < 10:
        month = '0' + str(month)
    else:
        month = str(month)
    result = []
    for i in range(d1, d2 + 1):
        if i < 10:
            result.append(f'0{i}.{month}.{year}')
        else:
            result.append(f'{i}.{month}.{year}')
    return result

dates = date_iter(4, 9, 8, 2022)

sums = []
for date in dates:
    cur.execute(f"""SELECT Services.price
                FROM Rendered
                INNER JOIN Services ON Rendered.service_id = Services.id
                INNER JOIN Clients ON Rendered.client_id = Clients.id
                WHERE Rendered.date = (?) AND (Clients.district = 'Центральный' OR Clients.district = 'Речной')
    """, (date,))
    for i in cur.fetchall():
        sums.append(i[0])

print(sums)
print(sum(sums))