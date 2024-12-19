import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('./19.xlsx')

sheet = wb['Лист1']
try:
    wb['2']
except:
    wb.create_sheet('2')
    wb.save('./19.xlsx')

sheet2 = wb["2"]


columns = sheet.max_row

n = 0

s = Side(border_style="double")

for i in range(1, columns + 1):
    array = []
    for j in sheet[i][:5]:
        if j.value != None:
            array.append(j.value)
    for j in range(1, len(array)):
        sum_odd = 0
        sum_even = 0
        if array[j]%2 == 0:
            sum_even += array[j]
        else:
            sum_odd += array[j]
        if array[j] == array[j-1]:
            n += 1
            break
        if j == 4:
            if sum_odd > sum_even:
                for x in sheet[i]:
                    x.fill = PatternFill("solid", fgColor="DDDDDD")
                    x.border = Border(s,s,s,s) 
                sheet[f"F{i}"] = array[0] + array[-1]
                sheet[f"G{i}"] = sum(array[1:3])
                sheet2.append(array)
                n += 1

wb.save('./19.xlsx')
print(n)