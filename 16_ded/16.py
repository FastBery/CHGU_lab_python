import sqlite3 as sq
import openpyxl
import datetime
import math

db = sq.connect('./16_ded/data_base/6.db')
cur = db.cursor()

cur.execute("PRAGMA foreign_keys=ON;") #making foreign_keys=on

cur.execute("""CREATE TABLE IF NOT EXISTS Tovari (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            zakupka INTEGER NOT NULL,
            prodazha INTEGER NOT NULL,
            kolichestvo INTEGER NOT NULL
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Postavki (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tovar_id TEXT NOT NULL,
            data TEXT NOT NULL,
            kolichestvo TEXT NOT NULL,
            FOREIGN KEY (tovar_id) REFERENCES Tovari (id) ON DELETE CASCADE
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Klienti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fio TEXT NOT NULL
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Magazini (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rayon TEXT NOT NULL
            )""")

cur.execute("""CREATE TABLE IF NOT EXISTS Operacii (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tovar_id INTEGER NOT NULL,
            klient_id INTEGER NOT NULL,
            magazin_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            tip TEXT NOT NULL,
            kolichestvo INTEGER NOT NULL,
            ostatok INTEGER NOT NULL,
            FOREIGN KEY (klient_id) REFERENCES Klienti (id) ON DELETE CASCADE,
            FOREIGN KEY (tovar_id) REFERENCES Tovari (id) ON DELETE CASCADE,
            FOREIGN KEY (magazin_id) REFERENCES Magazini (id) ON DELETE CASCADE
            )""")

wb = openpyxl.load_workbook('./16_ded/06.xlsx')

tovari_sheet = wb["Товары"]
postavki_sheet = wb["Поставки"]
operacii_sheet = wb["Операции"]
magazini_sheet = wb["Магазин"]
klienti_sheet = wb["Клиенты"]

#data for clients
tovari_data = []
for i in tovari_sheet.iter_rows(values_only=True, min_row=2):
    i = list(i)
    # print(type(i[2]))
    if i[2] != None:
        i[3] = round(i[2]*1.2)
        tovari_data.append(i)

# print(tovari_data[0:5])

try:
    cur.executemany("""INSERT INTO Tovari (id, name, zakupka, prodazha, kolichestvo) VALUES (?, ?, ?, ?, ?)""", tovari_data)
except:
    pass


#data for postavki
klienti_data = []
for i in klienti_sheet.iter_rows(values_only=True):
    klienti_data.append(i)

klienti_data.pop(0)

try:
    cur.executemany("""INSERT INTO Klienti (id, fio) VALUES (?, ?)""", klienti_data)
except:
    pass


#data for renders
postavki_data = []
for i in postavki_sheet.iter_rows(values_only=True, min_row=2):
    i = list(i)
    try:
        i[2] = i[2].strftime('%d/%m/%Y')
    except:
        pass
    if i[0] != None:
        postavki_data.append(i[0:4])

# print(postavki_data[0:5])

try:
    cur.executemany("""INSERT INTO Postavki (id, tovar_id, data, kolichestvo) VALUES (?, ?, ?, ?)""", postavki_data)
except:
    pass

magazini_data = []
for i in magazini_sheet.iter_rows(values_only=True):
    magazini_data.append(i)

magazini_data.pop(0)

try:
    cur.executemany("""INSERT INTO Magazini (id, rayon) VALUES (?, ?)""", magazini_data)
except:
    pass


operacii_data = []
for i in operacii_sheet.iter_rows(values_only=True, min_row=2):
    i = list(i)
    try:
        i[4] = i[4].strftime('%d/%m/%Y')
    except:
        pass
    if i[0] != None:
        operacii_data.append(i[0:8])

# print(operacii_data)

try:
    cur.executemany("""INSERT INTO Operacii (id, tovar_id, klient_id, magazin_id, data, tip, kolichestvo, ostatok) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""", operacii_data)
except:
    pass


cur.execute("""SELECT Operacii.kolichestvo, Operacii.data, Magazini.id
            FROM Operacii
            INNER JOIN Magazini ON Operacii.magazin_id = Magazini.id
            WHERE Operacii.tip = 'Покупка' AND magazini.rayon = 'Марьино'
""")

data = []
for i in cur.fetchall():
    data.append(i)

# print(data)

dictionary = dict()

for i in data:
    month = (i[1]).split('/')[1]
    if month in dictionary:
        dictionary[month] += i[0]
    else:
        dictionary[month] = i[0]
    
print(max(dictionary, key=dictionary.get))