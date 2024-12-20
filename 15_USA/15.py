import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('./20.xlsx')

sheet = wb.active
try:
    wb['2']
except:
    wb.create_sheet('2')
    wb.save('./20.xlsx')

sheet2 = wb["2"]


columns = sheet.max_row

n = 0

s = Side(border_style="double")

for i in range(1, columns + 1):
    array = []
    for j in sheet[i][:5]:
        if j.value != None:
            array.append(j.value)
    sorted_array = sorted(array)
    for j in range(1, len(array)):
        if sorted_array[j] == sorted_array[j-1]:
            break
        if j == 4:
            if (sorted_array[0] + sorted_array[-1]) < (sum(sorted_array[1:3])):
                for x in sheet[i]:
                    x.fill = PatternFill("solid", fgColor="DDDDDD")
                    x.border = Border(s,s,s,s) 
                # print(sorted_array, sorted_array[0] + sorted_array[-1], sum(sorted_array[1:3]))
                sheet[f"F{i}"] = sorted_array[0] + sorted_array[-1]
                sheet[f"G{i}"] = sum(sorted_array[1:3])
                sheet2.append(array)
                n += 1

wb.save('./20.xlsx')
print(n)